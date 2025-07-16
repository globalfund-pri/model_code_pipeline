from typing import Dict, Any, NamedTuple

import openpyxl
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import Reference, LineChart

from tgftools.analysis import PortfolioProjection
from tgftools.filehandler import Parameters
from tgftools.report import Report
from tgftools.utils import get_root_path, matmul


class SetOfPortfolioProjections(NamedTuple):
    IC: PortfolioProjection  # The main forward projection scenario for the investment case
    CF_InfAve: PortfolioProjection  # The counterfactual projection scenario to compute infections averted
    # (N.B., we may have multiple different ones for different diseases)
    CF_LivesSaved: PortfolioProjection  # The counterfactual projection scenario to compute lives saved
    # (N.B., we may have multiple different ones for different diseases)
    CF_LivesSaved_Malaria: pd.DataFrame  # The counterfactual for lives saved for malaria
    CF_InfectionsAverted_Malaria: pd.DataFrame  # The counterfactual for lives saved for malaria
    PARTNER: pd.DataFrame  # Dataframe containing partner data needed for reporting
    CF_forgraphs: pd.DataFrame  # Dataframe containing GP needed for reporting (N.B., may differ by disease)
    Info: Dict # Dictionary containing all the information on the ananlysis including files used and technical information


class STReport(Report):
    """This is the Report class. It accepts AnalysisResults for each disease and produces summary statistics.
    Each member function returns a Dict of the form {<label>: <stat>} which are assembled into an output Excel file.

    This is where we can access all data from all the scenarios we have defined, including the IC scenario and produce
    the data for the key graphs and key stats.

    CAUTION: the years and details of what is being extracted for the final report needs to be reviewed and done with
    care. The years and specific variables (e.g. tb deaths including or excluding hiv positive individuals).
    """

    def __init__(
            self,
            hiv: SetOfPortfolioProjections,
            tb: SetOfPortfolioProjections,
            malaria: SetOfPortfolioProjections,
            parameters: Parameters
    ):
        # Save arguments
        self.parameters = parameters
        self.hiv = hiv
        self.tb = tb
        self.malaria = malaria

    def info(self) -> pd.DataFrame:
        """Collate the information relating to the report, including which folders were being used and the details of
        the analysis (e.g., which approach was used (a or b), which funding envelope, how did we handle unalllocated
        amounts, did we adjust for innovation, etc). """
        return pd.DataFrame(
            data={
                'HIV': {**self.hiv.Info},
                'TB': {**self.tb.Info},
                'Malaria': {**self.malaria.Info}
            })

    def kpi1_and_kpi2(self) -> Dict[str, float]:
        """ Get the key stats for malaria.  """

        # HIV
        hiv_incidence_2021 = self.hiv.PARTNER["cases"].at[2021] / \
                             self.hiv.PARTNER["hivneg"].at[2021]
        hiv_incidence_2028 = self.hiv.IC.portfolio_results["cases"].at[2028, "model_central"] / \
                             self.hiv.IC.portfolio_results["hivneg"].at[2028, "model_central"]
        hiv_incidence_reduction_st = (hiv_incidence_2028 / hiv_incidence_2021 - 1) * 100


        hiv_mortality_2021 = self.hiv.PARTNER["deaths"].at[2021] / \
                             self.hiv.PARTNER["population"].at[2021]
        hiv_mortality_2028 = self.hiv.IC.portfolio_results["deaths"].at[2028, "model_central"] / \
                             self.hiv.IC.portfolio_results["population"].at[2028, "model_central"]
        hiv_mortality_reduction_st = (hiv_mortality_2028 / hiv_mortality_2021 - 1) * 100

        hiv_stats = {
                "HIV incidence in the year 2021": hiv_incidence_2021,
                "HIV incidence in the year 2028": hiv_incidence_2028,
                "Reduction in hiv incidence between the year 2028 compared to 2021": hiv_incidence_reduction_st,

                "HIV mortality rate in the year 2021": hiv_mortality_2021,
                "HIV mortality rate in the year 2028": hiv_mortality_2028,
                "Reduction in hiv mortality rate between the year 2028 compared to 2021": hiv_mortality_reduction_st,
        }

        # TB
        tb_incidence_2021 = self.tb.PARTNER["cases"].at[2021] / \
                            self.tb.PARTNER["population"].at[2021]
        tb_incidence_2028 = self.tb.IC.portfolio_results["cases"].at[2028, "model_central"] / \
                            self.tb.IC.portfolio_results["population"].at[2028, "model_central"]
        tb_incidence_reduction_st = (tb_incidence_2028 / tb_incidence_2021 - 1) * 100

        tb_mortality_2021 = self.tb.PARTNER["deaths"].at[2021] / \
                            self.tb.PARTNER["population"].at[2021]
        tb_mortality_2028 = self.tb.IC.portfolio_results["deaths"].at[2028, "model_central"] / \
                            self.tb.IC.portfolio_results["population"].at[2028, "model_central"]
        tb_mortality_reduction_st = (tb_mortality_2028 / tb_mortality_2021 - 1) * 100

        tb_mortality_hivneg_2021 = self.tb.PARTNER["deathshivneg"].at[2021] / \
                                   self.tb.PARTNER["population"].at[2021]
        tb_mortality_hivneg_2028 = self.tb.IC.portfolio_results["deathshivneg"].at[2028, "model_central"] / \
                                   self.tb.IC.portfolio_results["population"].at[2028, "model_central"]
        tb_mortality_hivneg_reduction_st = (tb_mortality_hivneg_2028 / tb_mortality_hivneg_2021 - 1) * 100

        tb_stats = {
            "TB incidence in the year 2021": tb_incidence_2021,
            "TB incidence in the year 2028": tb_incidence_2028,
            "Reduction in TB incidence between the year 2028 compared to 2021": tb_incidence_reduction_st,

            "TB mortality rate in the year 2021": tb_mortality_2021,
            "TB mortality rate in the year 2028": tb_mortality_2028,
            "Reduction in TB mortality rate between the year 2028 compared to 2021": tb_mortality_reduction_st,

            "TB mortality rate amongst hiv-negative individuals in the year 2021": tb_mortality_hivneg_2021,
            "TB mortality rate amongst hiv-negative individuals in the year 2028": tb_mortality_hivneg_2028,
            "Reduction in TB mortality rate amongst hiv-negative individuals between the year 2028 compared to 2021": tb_mortality_hivneg_reduction_st,
        }

        # Malaria
        malaria_incidence_2021 = self.malaria.PARTNER["cases"].at[2021] / \
                                 self.malaria.PARTNER["par"].at[2021]
        malaria_incidence_2028 = self.malaria.IC.portfolio_results["cases"].at[2028, "model_central"] / \
                                 self.malaria.IC.portfolio_results["par"].at[2028, "model_central"]
        malaria_incidence_reduction_st = (malaria_incidence_2028 / malaria_incidence_2021 - 1) * 100

        malaria_mortality_2021 = self.malaria.PARTNER["deaths"].at[2021] / \
                                 self.malaria.PARTNER["par"].at[2021]
        malaria_mortality_2028 = self.malaria.IC.portfolio_results["deaths"].at[2028, "model_central"] / \
                                 self.malaria.IC.portfolio_results["par"].at[2028, "model_central"]
        malaria_mortality_reduction_st = (malaria_mortality_2028 / malaria_mortality_2021 - 1) * 100

        malaria_stats ={
            "Malaria incidence in the year 2021": malaria_incidence_2021,
            "Malaria incidence in the year 2028": malaria_incidence_2028,
            "Reduction in malaria incidence between the year 2028 compared to 2021": malaria_incidence_reduction_st,

            "Malaria mortality rate in the year 2021": malaria_mortality_2021,
            "Malaria mortality rate in the year 2028": malaria_mortality_2028,
            "Reduction in malaria mortality rate between the year 2028 compared to 2021": malaria_mortality_reduction_st,
        }

        return {**hiv_stats , **tb_stats, **malaria_stats}

    def service_coverage(self) -> pd.DataFrame:
        """Dump into a pd.DataFrame all the country results, ready to be written to Excel worksheet. """

        list_of_dfs = list()
        for disease, obj in zip(('HIV', 'TB', 'MALARIA'), (self.hiv, self.tb, self.malaria)):
            # Get the service coverage indicators that will be reported for this disease
            all_indicators = self.parameters.get_indicators_for(disease)
            service_cov_indicators = all_indicators.loc[all_indicators['is_service_coverage_indicator'] == True, 'description'].to_dict()

            for country, results in obj.IC.country_results.items():
                for indicator, description in service_cov_indicators.items():
                    try:
                        df = results.model_projection[indicator].loc[range(2024,2029)]  # reporting in range 2024,2028 inclusive.
                        df['disease'] = disease
                        df['country'] = country
                        df['indicator'] = indicator
                        df['description'] = description
                        list_of_dfs.append(df)
                    except KeyError:
                        print(f"Could not find model results: {country=} {indicator=} {disease=} ")

        order_of_columns = [
            'disease', 'country', 'year', 'indicator', 'description', 'model_low', 'model_central', 'model_high',
            'pf_low', 'pf_central', 'pf_high', 'partner_low', 'partner_central', 'partner_high',
        ]
        sheet = pd.concat(list_of_dfs).reset_index()[order_of_columns].sort_values(['disease', 'country', 'year', 'indicator'])

        # Add a 'check' column to signify discrepancies
        tol = 0.1
        model_range_does_not_include_partner_estimate = (
                    sheet['partner_central'].notnull() & ~sheet['partner_central'].between(
                sheet['model_low'] * (1. - tol),
                sheet['model_high'] * (1. + tol),
                inclusive='both'
            ))

        model_range_does_not_include_pf_estimate = (
            sheet['pf_central'].notnull() & ~sheet['pf_central'].between(
                sheet['model_low'] * (1. - tol),
                sheet['model_high'] * (1. + tol),
                inclusive='both'
            )
        )

        sheet['est_out_of_range'] = model_range_does_not_include_partner_estimate | model_range_does_not_include_pf_estimate

        return sheet

    def _post_processing_on_workbook(self, workbook: Workbook):
        """Add custom formatting to workbook."""

        # Color rows based on 'est_out_of_range' column value: red if 'est_out_of_range' column is true, meaning some problem.

        sheet = workbook['service_co']

        # Get all rows and columns in the worksheet
        max_row = sheet.max_row
        max_col = sheet.max_column

        # Find the 'est_out_of_range' column index
        check_col = None
        for col in range(1, max_col + 1):
            if sheet.cell(row=1, column=col).value == 'est_out_of_range':
                check_col = col
                break

        if check_col:
            for row in range(2, max_row + 1):  # Start from row 2 to skip header
                check_value = sheet.cell(row=row, column=check_col).value
                if check_value == True:  # If check is True, colour the row red
                    for col in range(1, max_col + 1):
                        cell = sheet.cell(row=row, column=col)
                        cell.fill = openpyxl.styles.PatternFill(
                            start_color='FFCCCC',
                            end_color='FFCCCC',
                            fill_type='solid'
                        )
