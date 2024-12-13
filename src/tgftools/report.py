from pathlib import Path
from pprint import pprint
from typing import Optional, Dict

import pandas as pd

from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook

from tgftools.utils import current_date_and_time_as_string, get_commit_revision_number


class Report:
    """This is the BaseClass for Report classes. It provides the core functionality to generate reports. It can be
    inherited from to allow it to accept sets of PortfolioProjections for the diseases. It intended that each member
    function will either: (i) return a Dict of the form {<label>: <stat>}, or (ii) return a pd.DataFrame. These can be
    assembled into an output Excel file: contents of dicts are written to the 'Stats' worksheet; contents of
    pd.DataFrames are written to their own sheet of the same name.
    """

    def __init__(self, *args, **kwargs):
        """Initialise the Report Class"""

    def _get_all_funcs_to_generate_stats(self) -> list[str]:
        """Returns a list of the functions in the class that will generate statistics (i.e., any function with a name
        that does not start with "_" and is not called "report".
        """
        return sorted(
            [
                name
                for name in dir(self)
                if (
                    not name.startswith("_")
                    and (not name.startswith("report"))
                    and callable(self.__getattribute__(name))
            )
            ]
        )

    def report(self, filename: Optional[Path] = None) -> Dict:
        """Run all member functions, print the results to screen, returns the results in the form of dictionary and
        (if filename provided) assemble them into an Excel file and draw graphs."""

        all_results_for_stats_pages = dict()  # Storage for all the results
        all_results_for_individual_worksheets = dict()

        all_funcs = self._get_all_funcs_to_generate_stats()
        for ch_name in all_funcs:
            pprint(f"** {ch_name} **")
            output = self.__getattribute__(ch_name)()
            pprint(output)

            if isinstance(output, dict):
                all_results_for_stats_pages[ch_name] = output
            elif isinstance(output, pd.DataFrame):
                all_results_for_individual_worksheets[ch_name] = output
            else:
                raise ValueError(f"Return from {ch_name} function is not of recognised type ({type(ch_name)}).")

        # Compile the results for the 'stats' summary
        results_for_main = list()
        for func_name, func_results in all_results_for_stats_pages.items():
            for stat_name, stat_result in func_results.items():
                results_for_main.append([func_name, stat_name, stat_result])

        if filename is not None:
            # Write to Excel
            wb = Workbook()

            # Write to 'info' sheet
            work_sheet_info = wb.active
            work_sheet_info.title = 'git'
            work_sheet_info.append(['date-time stamp', current_date_and_time_as_string()])
            work_sheet_info.append(['commit', get_commit_revision_number()])

            # Write to 'stats' worksheet:
            work_sheet_stats = wb.create_sheet()
            work_sheet_stats.title = 'stats'
            for line in results_for_main:
                work_sheet_stats.append(line)

            # Write results to 'individual' worksheet
            for func_name, func_results in all_results_for_individual_worksheets.items():
                work_sheet = wb.create_sheet()
                work_sheet.title = func_name[0:10]  # truncate to first ten characters, as requirement of Excel
                for r in dataframe_to_rows(func_results.reset_index(), index=False, header=True):
                    work_sheet.append(r)

            # Do any post-processing that may be required
            self._post_processing_on_workbook(wb)

            # Save
            wb.save(filename)

        return {
            # Returning in the same format as the Excel file:
            # * key='main': a pd.DataFrame contains all the scalar stats from individual functions
            # * all other keys/sheets: pd.DataFrames from all the functions that returned pd.DataFrames
            'stats': (
                pd.DataFrame(results_for_main)
                .rename(columns={0: 'Function', 1: 'Key', 2: 'Value'})
            ),
            **all_results_for_individual_worksheets,
        }

    def _post_processing_on_workbook(self, workbook: Workbook):
        """Do anything necessary to post-process the workbook: for instance, create graphs on certain worksheets."""
        pass
